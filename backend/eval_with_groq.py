#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════════════╗
║   NormalizerDB — GROQ Evaluation Suite                              ║
║   Tests normalization quality using LLM-as-judge (Groq API)        ║
╚══════════════════════════════════════════════════════════════════════╝

USAGE:
  # Single file
  python eval_with_groq.py --input orders.xlsx

  # Batch: entire folder
  python eval_with_groq.py --folder ./test_files/

  # Set API key inline (or put it in .env / environment)
  GROQ_API_KEY=gsk_... python eval_with_groq.py --folder ./test_files/

  # Use a specific Groq model
  python eval_with_groq.py --folder ./test_files/ --model llama3-70b-8192

INSTALL:
  pip install groq pandas openpyxl python-dotenv rich

OUTPUT:
  • Live colored terminal report (via Rich)
  • eval_results/  folder with per-file JSON + final summary Excel
"""

import os
import sys
import json
import time
import argparse
import textwrap
import traceback
import tempfile
from pathlib import Path
from datetime import datetime

import pandas as pd

# ── Optional rich terminal output (graceful fallback if not installed) ──
try:
    from rich.console import Console
    from rich.table import Table
    from rich.panel import Panel
    from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn
    from rich.syntax import Syntax
    from rich.rule import Rule
    from rich import box
    RICH = True
    console = Console()
except ImportError:
    RICH = False
    class _FallbackConsole:
        def print(self, *a, **kw):
            txt = " ".join(str(x) for x in a)
            # Strip rich markup tags
            import re
            print(re.sub(r'\[.*?\]', '', txt))
        def rule(self, title=""):   print(f"\n{'─'*60} {title}")
        def log(self, *a, **kw):    self.print(*a)
    console = _FallbackConsole()

# ── Optional python-dotenv ───────────────────────────────────────────────
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# ── Groq client ─────────────────────────────────────────────────────────
try:
    from groq import Groq
except ImportError:
    console.print("[bold red]✗ groq package not installed.[/bold red]  Run:  pip install groq")
    sys.exit(1)

# ── Local engine (same folder) ───────────────────────────────────────────
sys.path.insert(0, str(Path(__file__).parent))
try:
    import normalize_engine as engine
except ImportError as e:
    console.print(f"[bold red]✗ Cannot import normalize_engine: {e}[/bold red]")
    sys.exit(1)


# ════════════════════════════════════════════════════════════════════════
#  CONFIG
# ════════════════════════════════════════════════════════════════════════

DEFAULT_MODEL   = "llama-3.3-70b-versatile"
OUTPUT_DIR      = Path("eval_results")
MAX_ROWS_SAMPLE = 8     # rows to include in prompt (keep tokens low)
MAX_COLS_SAMPLE = 20    # cols to include

NF_CRITERIA = {
    "NF1": "All cells are atomic (no comma-separated lists). No repeating groups. No composite attributes.",
    "NF2": "Every non-key attribute depends on the WHOLE primary key (no partial dependencies).",
    "NF3": "No non-key attribute transitively depends on the primary key. No stored derived/calculated fields.",
    "BCNF": "For every non-trivial FD X→Y, X is a superkey.",
    "NF4": "No independent multi-valued dependencies. Each independent fact is in its own table.",
    "NF5": "No join dependencies remain unless implied by candidate keys. Lossless decomposition.",
}

SCORE_RUBRIC = """
Score each criterion 0–10:
  10 = Perfectly satisfied, nothing to improve
   8 = Mostly correct, minor issues
   6 = Partially correct, notable gaps
   4 = Attempted but significant violations remain
   2 = Barely addressed
   0 = Completely ignored or made worse

Also give an overall score 0–100 (weighted average of NF scores × their weights):
  NF1: 20 points, NF2: 20 points, NF3: 15 points,
  BCNF: 15 points, NF4: 15 points, NF5: 15 points.
"""


# ════════════════════════════════════════════════════════════════════════
#  HELPERS
# ════════════════════════════════════════════════════════════════════════

def df_to_text_table(df: pd.DataFrame, max_rows=MAX_ROWS_SAMPLE, max_cols=MAX_COLS_SAMPLE) -> str:
    """Render a DataFrame as a compact markdown-ish text table."""
    sub = df.iloc[:max_rows, :max_cols].fillna("NULL").astype(str)
    cols = list(sub.columns)
    widths = [max(len(c), sub[c].str.len().max()) + 1 for c in cols]
    widths = [min(w, 25) for w in widths]  # cap width

    def row_str(vals):
        return "│ " + " │ ".join(str(v)[:w].ljust(w) for v, w in zip(vals, widths)) + " │"

    sep = "├─" + "─┼─".join("─" * w for w in widths) + "─┤"
    header = "│ " + " │ ".join(c[:w].ljust(w) for c, w in zip(cols, widths)) + " │"

    lines = ["┌─" + "─┬─".join("─" * w for w in widths) + "─┐",
             header, sep]
    for _, r in sub.iterrows():
        lines.append(row_str(list(r)))
    lines.append("└─" + "─┴─".join("─" * w for w in widths) + "─┘")

    if len(df) > max_rows:
        lines.append(f"  ... {len(df) - max_rows} more rows not shown")
    if len(df.columns) > max_cols:
        lines.append(f"  ... {len(df.columns) - max_cols} more columns not shown")
    return "\n".join(lines)


def schema_summary(tables: dict) -> str:
    """Compact text representation of the normalized schema."""
    lines = []
    for tname, (df, pk, fk) in tables.items():
        pk_str = ", ".join(pk) if pk else "—"
        fk_str = ", ".join(fk) if fk else "none"
        lines.append(f"\n  ┌── Table: {tname}")
        lines.append(f"  │  Rows: {len(df)}  Cols: {len(df.columns)}")
        lines.append(f"  │  PK  : {pk_str}")
        lines.append(f"  │  FK  : {fk_str}")
        lines.append(f"  │  Columns: {', '.join(df.columns)}")
        # Sample data
        sample = df_to_text_table(df, max_rows=3, max_cols=10)
        for line in sample.split("\n"):
            lines.append(f"  │    {line}")
        lines.append(f"  └{'─'*50}")
    return "\n".join(lines)


def build_prompt(original_df: pd.DataFrame, tables: dict, sql_snippet: str,
                 filename: str, nf_log: dict) -> str:
    """
    Build the evaluation prompt sent to Groq.
    Structured to get consistent JSON back.
    """
    orig_sample  = df_to_text_table(original_df)
    schema_text  = schema_summary(tables)
    log_text     = "\n".join(
        f"  [{stage}]" + "\n".join(f"    {m}" for m in msgs)
        for stage, msgs in nf_log.items()
    )
    nf_criteria_text = "\n".join(f"  • {k}: {v}" for k, v in NF_CRITERIA.items())

    prompt = f"""
You are an expert database engineer and academic evaluator specializing in relational database normalization theory (NF1 through NF5).

You have been given:
1. An ORIGINAL denormalized Excel table (filename: {filename})
2. The NORMALIZED output produced by an automated normalization engine
3. The normalization log showing what transformations were applied
4. A SQL schema snippet

Your task is to evaluate the quality of the normalization and return a structured JSON score.

═══════════════════════════════════════════════
ORIGINAL TABLE SAMPLE ({len(original_df)} rows × {len(original_df.columns)} cols)
═══════════════════════════════════════════════
{orig_sample}

═══════════════════════════════════════════════
NORMALIZATION LOG
═══════════════════════════════════════════════
{log_text}

═══════════════════════════════════════════════
NORMALIZED SCHEMA ({len(tables)} tables)
═══════════════════════════════════════════════
{schema_text}

═══════════════════════════════════════════════
SQL SCHEMA SNIPPET (first 60 lines)
═══════════════════════════════════════════════
{sql_snippet}

═══════════════════════════════════════════════
NORMAL FORM CRITERIA TO EVALUATE
═══════════════════════════════════════════════
{nf_criteria_text}

═══════════════════════════════════════════════
SCORING RUBRIC
═══════════════════════════════════════════════
{SCORE_RUBRIC}

═══════════════════════════════════════════════
RESPONSE FORMAT — respond ONLY with valid JSON, no markdown fences, no preamble
═══════════════════════════════════════════════
{{
  "file": "{filename}",
  "overall_score": <0-100 integer>,
  "overall_verdict": "<one of: EXCELLENT | GOOD | ACCEPTABLE | NEEDS_WORK | POOR>",
  "nf_scores": {{
    "NF1": {{ "score": <0-10>, "reasoning": "<1-2 sentences>", "issues": ["<issue1>", ...] }},
    "NF2": {{ "score": <0-10>, "reasoning": "<1-2 sentences>", "issues": [] }},
    "NF3": {{ "score": <0-10>, "reasoning": "<1-2 sentences>", "issues": [] }},
    "BCNF": {{ "score": <0-10>, "reasoning": "<1-2 sentences>", "issues": [] }},
    "NF4": {{ "score": <0-10>, "reasoning": "<1-2 sentences>", "issues": [] }},
    "NF5": {{ "score": <0-10>, "reasoning": "<1-2 sentences>", "issues": [] }}
  }},
  "strengths": ["<strength1>", "<strength2>"],
  "weaknesses": ["<weakness1>", "<weakness2>"],
  "recommendations": ["<rec1>", "<rec2>"],
  "is_valid_sql": <true|false>,
  "sql_issues": ["<sql_issue1>"]
}}
""".strip()
    return prompt


def parse_groq_response(raw: str) -> dict:
    """
    Extract JSON from Groq's response — handles markdown fences, preamble, etc.
    """
    raw = raw.strip()

    # Strip markdown code fences
    import re
    raw = re.sub(r'^```(?:json)?\s*', '', raw, flags=re.MULTILINE)
    raw = re.sub(r'\s*```$', '', raw, flags=re.MULTILINE)
    raw = raw.strip()

    # Find first { ... } block
    start = raw.find('{')
    end   = raw.rfind('}')
    if start == -1 or end == -1:
        raise ValueError(f"No JSON object found in response:\n{raw[:300]}")

    json_str = raw[start:end+1]
    return json.loads(json_str)


# ════════════════════════════════════════════════════════════════════════
#  CORE EVALUATOR
# ════════════════════════════════════════════════════════════════════════

class NormalizationEvaluator:
    def __init__(self, api_key: str, model: str = DEFAULT_MODEL):
        self.client = Groq(api_key=api_key)
        self.model  = model
        OUTPUT_DIR.mkdir(exist_ok=True)

    def evaluate_file(self, excel_path: str) -> dict:
        """
        Run full pipeline on one Excel file, then ask Groq to score the result.
        Returns the evaluation dict.
        """
        fname = Path(excel_path).name
        console.print(f"\n[bold cyan]▶ Processing:[/bold cyan] {fname}")

        result = {
            "file":       fname,
            "path":       str(excel_path),
            "timestamp":  datetime.now().isoformat(),
            "status":     "error",
            "engine_ok":  False,
            "groq_ok":    False,
            "error":      None,
            "table_count": 0,
            "overall_score": 0,
            "nf_scores": {},
        }

        # ── Step 1: Run normalization engine ─────────────────────
        tmp_xlsx = str(OUTPUT_DIR / f"_tmp_{Path(excel_path).stem}_norm.xlsx")
        tmp_sql  = str(OUTPUT_DIR / f"_tmp_{Path(excel_path).stem}_schema.sql")

        try:
            console.print("  [dim]→ Running normalization engine…[/dim]")
            original_df = pd.read_excel(excel_path, sheet_name=0)

            tables, nf_log = engine.normalize(
                input_excel=excel_path,
                output_excel=tmp_xlsx,
                output_sql=tmp_sql,
                sheet_name=0
            )
            result["engine_ok"]   = True
            result["table_count"] = len(tables)
            console.print(f"  [green]✓ Engine done:[/green] {len(tables)} tables produced")

        except Exception as e:
            result["error"] = f"Engine error: {e}\n{traceback.format_exc()}"
            console.print(f"  [red]✗ Engine failed: {e}[/red]")
            return result

        # ── Step 2: Read SQL snippet ──────────────────────────────
        try:
            with open(tmp_sql, encoding="utf-8") as f:
                sql_lines = f.readlines()
            sql_snippet = "".join(sql_lines[:60])
        except Exception:
            sql_snippet = "(SQL file could not be read)"

        # ── Step 3: Build prompt and call Groq ────────────────────
        console.print("  [dim]→ Sending to Groq for evaluation…[/dim]")
        prompt = build_prompt(original_df, tables, sql_snippet, fname, nf_log)

        try:
            response = self.client.chat.completions.create(
                model=self.model,
                messages=[
                    {
                        "role": "system",
                        "content": (
                            "You are a strict, expert database normalization evaluator. "
                            "You ALWAYS respond with valid JSON only — no markdown, no extra text. "
                            "Be precise and rigorous in your assessment."
                        )
                    },
                    {"role": "user", "content": prompt}
                ],
                temperature=0.1,   # low temp → deterministic scoring
                max_tokens=2000,
            )

            raw_reply = response.choices[0].message.content
            evaluation = parse_groq_response(raw_reply)
            result.update(evaluation)
            result["groq_ok"] = True
            result["status"]  = "done"
            result["raw_groq_response"] = raw_reply

            score = evaluation.get("overall_score", 0)
            verdict = evaluation.get("overall_verdict", "?")
            console.print(
                f"  [bold green]✓ Groq score: {score}/100[/bold green] "
                f"[yellow]{verdict}[/yellow]"
            )

        except Exception as e:
            result["error"] = f"Groq error: {e}\n{traceback.format_exc()}"
            console.print(f"  [red]✗ Groq evaluation failed: {e}[/red]")

        # ── Step 4: Save per-file JSON ────────────────────────────
        json_path = OUTPUT_DIR / f"{Path(excel_path).stem}_eval.json"
        with open(json_path, "w", encoding="utf-8") as f:
            # Exclude raw response from JSON to keep it clean
            clean = {k: v for k, v in result.items() if k != "raw_groq_response"}
            json.dump(clean, f, indent=2, ensure_ascii=False)
        console.print(f"  [dim]→ Saved: {json_path}[/dim]")

        return result

    def run_batch(self, paths: list[str]) -> list[dict]:
        """Evaluate multiple Excel files and produce a summary report."""
        console.print(Panel(
            f"[bold]NormalizerDB — GROQ Evaluation Suite[/bold]\n"
            f"Model : [cyan]{self.model}[/cyan]\n"
            f"Files : [yellow]{len(paths)}[/yellow]",
            title="⚡ Starting Batch Evaluation",
            border_style="blue",
        ))

        results = []
        for i, path in enumerate(paths, 1):
            console.rule(f"[{i}/{len(paths)}] {Path(path).name}")
            res = self.evaluate_file(path)
            results.append(res)
            # Brief pause between calls to respect rate limits
            if i < len(paths):
                time.sleep(1.5)

        self._print_summary(results)
        self._save_summary_excel(results)
        return results

    # ── Pretty terminal summary ───────────────────────────────────────

    def _print_summary(self, results: list[dict]):
        console.rule("[bold yellow]EVALUATION SUMMARY[/bold yellow]")

        if RICH:
            table = Table(box=box.ROUNDED, show_header=True, header_style="bold cyan")
            table.add_column("File",         style="white",   max_width=28)
            table.add_column("Tables",       style="cyan",    justify="center")
            table.add_column("Score",        style="bold",    justify="center")
            table.add_column("Verdict",                       justify="center")
            table.add_column("NF1", justify="center")
            table.add_column("NF2", justify="center")
            table.add_column("NF3", justify="center")
            table.add_column("BCNF", justify="center")
            table.add_column("NF4", justify="center")
            table.add_column("NF5", justify="center")
            table.add_column("Issues", style="dim", max_width=30)

            for r in results:
                nf = r.get("nf_scores", {})
                score = r.get("overall_score", 0)
                verdict = r.get("overall_verdict", "ERROR" if r["status"]=="error" else "?")

                score_color = (
                    "green" if score >= 80 else
                    "yellow" if score >= 60 else
                    "red"
                )
                verdict_color = {
                    "EXCELLENT": "bold green", "GOOD": "green",
                    "ACCEPTABLE": "yellow",    "NEEDS_WORK": "red",
                    "POOR": "bold red",
                }.get(verdict, "dim")

                def nf_cell(key):
                    s = nf.get(key, {}).get("score", "–")
                    if s == "–": return "–"
                    c = "green" if s >= 8 else "yellow" if s >= 5 else "red"
                    return f"[{c}]{s}[/{c}]"

                all_issues = []
                for nf_key, nf_data in nf.items():
                    all_issues += nf_data.get("issues", [])
                issues_str = "; ".join(all_issues[:2]) if all_issues else "none"

                table.add_row(
                    r["file"][:28],
                    str(r.get("table_count", "?")),
                    f"[{score_color}]{score}/100[/{score_color}]",
                    f"[{verdict_color}]{verdict}[/{verdict_color}]",
                    nf_cell("NF1"), nf_cell("NF2"), nf_cell("NF3"),
                    nf_cell("BCNF"), nf_cell("NF4"), nf_cell("NF5"),
                    issues_str[:30],
                )

            console.print(table)
        else:
            # Plain text fallback
            print(f"\n{'File':<28} {'Score':>7} {'Verdict':<12} NF1 NF2 NF3 BCN NF4 NF5")
            print("─" * 80)
            for r in results:
                nf = r.get("nf_scores", {})
                def s(k): return str(nf.get(k, {}).get("score", "?"))
                print(
                    f"{r['file'][:28]:<28} "
                    f"{r.get('overall_score', 0):>6}/100 "
                    f"{r.get('overall_verdict','?'):<12} "
                    f"{s('NF1'):>3} {s('NF2'):>3} {s('NF3'):>3} "
                    f"{s('BCNF'):>3} {s('NF4'):>3} {s('NF5'):>3}"
                )

        # Aggregate stats
        done   = [r for r in results if r["status"] == "done"]
        failed = [r for r in results if r["status"] != "done"]
        if done:
            avg = sum(r.get("overall_score", 0) for r in done) / len(done)
            console.print(
                f"\n  [bold]Total:[/bold] {len(results)} files  "
                f"[green]{len(done)} succeeded[/green]  "
                f"[red]{len(failed)} failed[/red]  "
                f"[bold cyan]Average score: {avg:.1f}/100[/bold cyan]"
            )

        # Detailed breakdown per file
        for r in results:
            if r["status"] != "done":
                continue
            console.rule(f"[dim]{r['file']}[/dim]")
            nf = r.get("nf_scores", {})
            for nf_key, nf_data in nf.items():
                score = nf_data.get("score", "?")
                reasoning = nf_data.get("reasoning", "")
                issues = nf_data.get("issues", [])
                score_tag = (
                    f"[green]{score}/10[/green]" if isinstance(score, int) and score >= 8 else
                    f"[yellow]{score}/10[/yellow]" if isinstance(score, int) and score >= 5 else
                    f"[red]{score}/10[/red]"
                )
                console.print(f"  {nf_key:5s} {score_tag}  {reasoning}")
                for issue in issues:
                    console.print(f"         [dim]⚠ {issue}[/dim]")

            strengths = r.get("strengths", [])
            weaknesses = r.get("weaknesses", [])
            recs = r.get("recommendations", [])

            if strengths:
                console.print(f"  [green]Strengths:[/green] {'; '.join(strengths[:2])}")
            if weaknesses:
                console.print(f"  [red]Weaknesses:[/red] {'; '.join(weaknesses[:2])}")
            if recs:
                console.print(f"  [cyan]Recommendations:[/cyan] {'; '.join(recs[:2])}")

    # ── Save summary to Excel ─────────────────────────────────────────

    def _save_summary_excel(self, results: list[dict]):
        """Write a nicely formatted summary Excel workbook."""
        try:
            from openpyxl import Workbook as WB
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = WB()
            ws = wb.active
            ws.title = "Evaluation Summary"

            border = Border(
                left=Side(style='thin'),   right=Side(style='thin'),
                top=Side(style='thin'),    bottom=Side(style='thin'),
            )

            headers = [
                "File", "Status", "Tables", "Score/100", "Verdict",
                "NF1/10", "NF2/10", "NF3/10", "BCNF/10", "NF4/10", "NF5/10",
                "NF1 Issues", "Strengths", "Weaknesses", "Recommendations",
                "SQL Valid", "Error"
            ]

            hdr_fill = PatternFill("solid", start_color="1B2631")
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
                cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal='center', wrap_text=True)
                cell.border = border

            score_fills = {
                "green":  "D5F5E3",
                "yellow": "FEF9E7",
                "red":    "FADBD8",
            }

            for ri, r in enumerate(results, 2):
                nf = r.get("nf_scores", {})
                score = r.get("overall_score", 0)
                verdict = r.get("overall_verdict", r.get("status", "?").upper())

                def nf_score(k): return nf.get(k, {}).get("score", "?")
                def nf_issues(k): return "; ".join(nf.get(k, {}).get("issues", []))

                row_data = [
                    r["file"],
                    r["status"],
                    r.get("table_count", "?"),
                    score,
                    verdict,
                    nf_score("NF1"), nf_score("NF2"), nf_score("NF3"),
                    nf_score("BCNF"), nf_score("NF4"), nf_score("NF5"),
                    nf_issues("NF1"),
                    "; ".join(r.get("strengths", [])[:2]),
                    "; ".join(r.get("weaknesses", [])[:2]),
                    "; ".join(r.get("recommendations", [])[:2]),
                    "Yes" if r.get("is_valid_sql") else "No",
                    str(r.get("error") or ""),
                ]

                fill_color = (
                    score_fills["green"]  if isinstance(score, int) and score >= 80 else
                    score_fills["yellow"] if isinstance(score, int) and score >= 60 else
                    score_fills["red"]
                )

                for ci, val in enumerate(row_data, 1):
                    cell = ws.cell(row=ri, column=ci, value=val)
                    cell.font = Font(name="Arial", size=10)
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    cell.border = border
                    # Color code the score column
                    if ci == 4:
                        cell.fill = PatternFill("solid", start_color=fill_color)
                        cell.font = Font(bold=True, name="Arial", size=10)

            # Auto column widths
            col_widths = [28, 10, 8, 10, 12, 8, 8, 8, 8, 8, 8, 30, 35, 35, 40, 10, 40]
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

            ws.row_dimensions[1].height = 30
            ws.freeze_panes = "A2"

            # Add per-NF detail sheet
            ws2 = wb.create_sheet("NF Detail")
            detail_headers = ["File", "NF", "Score/10", "Reasoning", "Issues"]
            for ci, h in enumerate(detail_headers, 1):
                cell = ws2.cell(row=1, column=ci, value=h)
                cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
                cell.fill = hdr_fill
                cell.border = border

            dr = 2
            for r in results:
                nf = r.get("nf_scores", {})
                for nf_key, nf_data in nf.items():
                    s = nf_data.get("score", "?")
                    fill_c = (
                        "D5F5E3" if isinstance(s, int) and s >= 8 else
                        "FEF9E7" if isinstance(s, int) and s >= 5 else
                        "FADBD8"
                    )
                    for ci, val in enumerate([
                        r["file"], nf_key, s,
                        nf_data.get("reasoning", ""),
                        "; ".join(nf_data.get("issues", []))
                    ], 1):
                        cell = ws2.cell(row=dr, column=ci, value=val)
                        cell.font = Font(name="Arial", size=10)
                        cell.alignment = Alignment(wrap_text=True)
                        cell.border = border
                        if ci == 3:
                            cell.fill = PatternFill("solid", start_color=fill_c)
                    dr += 1

            for ci, w in enumerate([28, 8, 10, 50, 50], 1):
                ws2.column_dimensions[get_column_letter(ci)].width = w

            out_path = OUTPUT_DIR / f"eval_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(out_path)
            console.print(f"\n  [bold green]📊 Summary saved:[/bold green] {out_path}")

        except Exception as e:
            console.print(f"  [yellow]⚠ Could not save Excel summary: {e}[/yellow]")


# ════════════════════════════════════════════════════════════════════════
#  CLI ENTRY POINT
# ════════════════════════════════════════════════════════════════════════

def collect_excel_files(folder: str) -> list[str]:
    """Recursively find all .xlsx and .xls files in a folder."""
    folder = Path(folder)
    files = sorted(folder.glob("**/*.xlsx")) + sorted(folder.glob("**/*.xls"))
    # Skip temp files created by the engine
    return [str(f) for f in files if not f.name.startswith("_tmp_")]


def main():
    parser = argparse.ArgumentParser(
        description="NormalizerDB GROQ Evaluation Suite",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=textwrap.dedent("""
            Examples:
              python eval_with_groq.py --input orders.xlsx
              python eval_with_groq.py --folder ./test_data/
              python eval_with_groq.py --folder ./test_data/ --model mixtral-8x7b-32768
              GROQ_API_KEY=gsk_... python eval_with_groq.py --input data.xlsx

            Available Groq models:
              llama-3.3-70b-versatile      (default — best reasoning)
              llama3-8b-8192       (faster, cheaper)
              mixtral-8x7b-32768   (large context window)
              gemma2-9b-it         (alternative)
        """)
    )
    parser.add_argument("--input",    type=str, help="Single Excel file to evaluate")
    parser.add_argument("--folder",   type=str, help="Folder of Excel files (batch mode)")
    parser.add_argument("--model",    type=str, default=DEFAULT_MODEL,
                        help=f"Groq model to use (default: {DEFAULT_MODEL})")
    parser.add_argument("--api-key",  type=str, default=None,
                        help="Groq API key (or set GROQ_API_KEY env var)")
    parser.add_argument("--out-dir",  type=str, default="eval_results",
                        help="Output directory for results (default: eval_results/)")
    args = parser.parse_args()

    # Resolve output dir
    global OUTPUT_DIR
    OUTPUT_DIR = Path(args.out_dir)

    # Resolve API key
    api_key = (
        args.api_key
        or os.environ.get("GROQ_API_KEY")
    )
    if not api_key:
        console.print(
            "[bold red]✗ No Groq API key found.[/bold red]\n"
            "  Set it via:  export GROQ_API_KEY=gsk_...\n"
            "  Or pass:     --api-key gsk_...\n"
            "  Or create a .env file with:  GROQ_API_KEY=gsk_..."
        )
        sys.exit(1)

    # Collect files
    if args.input:
        files = [args.input]
    elif args.folder:
        files = collect_excel_files(args.folder)
        if not files:
            console.print(f"[red]✗ No Excel files found in '{args.folder}'[/red]")
            sys.exit(1)
    else:
        parser.print_help()
        sys.exit(0)

    console.print(f"  Found [yellow]{len(files)}[/yellow] file(s) to evaluate")
    for f in files:
        console.print(f"    • {Path(f).name}")

    # Run
    evaluator = NormalizationEvaluator(api_key=api_key, model=args.model)
    results = evaluator.run_batch(files)

    # Exit code: 0 if all passed (score ≥ 60), 1 otherwise
    all_passed = all(r.get("overall_score", 0) >= 60 for r in results if r["status"] == "done")
    sys.exit(0 if all_passed else 1)


if __name__ == "__main__":
    main()
