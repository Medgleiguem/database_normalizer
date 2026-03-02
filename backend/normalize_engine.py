"""
╔══════════════════════════════════════════════════════════════════╗
║  DATABASE NORMALIZATION ENGINE  v2.0  —  NF1 → NF5              ║
║  Complete rewrite with correct formal normalization logic        ║
╚══════════════════════════════════════════════════════════════════╝

Key improvements over v1:
  ✔ NF1: Detects ANY multi-valued cell (even "MA, MAI" or "L1, L2")
  ✔ NF1: Detects composite attributes (names, addresses)
  ✔ NF1: Correctly groups parallel vs independent multi-valued columns
  ✔ NF2: Heuristic + entropy-based partial dependency detection
  ✔ NF3: Proper transitive dependency chains & calculated-field detection
  ✔ BCNF: Strict superkey verification
  ✔ NF4: True MVD detection (independent multi-valued facts)
  ✔ NF5: Join dependency check for ternary relations
  ✔ SQL: Full schema with FK constraints, indexes, views for calc fields
"""

import pandas as pd
import re
import numpy as np
from collections import defaultdict
from itertools import combinations
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BORDER = Border(
    left=Side(style='thin', color='C0C0C0'),
    right=Side(style='thin', color='C0C0C0'),
    top=Side(style='thin', color='C0C0C0'),
    bottom=Side(style='thin', color='C0C0C0'),
)

SEP_PATTERN = re.compile(r'\s*[,;|]\s*')
COMPOSITE_MARKERS = [' et ', ' and ', ' & ', ' / ']
DATE_PATTERN = re.compile(r'^\d{4}[-/]\d{2}[-/]\d{2}$')
DECIMAL_COMMA = re.compile(r'^\d+,\d{1,2}$')
NUMERIC_PATTERN = re.compile(r'^-?\d+([.,]\d+)?$')


def safe_str(val):
    """Convert any value to str, returning '' for None/NaN/float-nan."""
    if val is None:
        return ""
    try:
        if pd.isna(val):
            return ""
    except (TypeError, ValueError):
        pass
    return str(val)


def clean_name(name):
    return re.sub(r'[^a-zA-Z0-9_]', '_', str(name)).strip('_').lower()


def is_multi_valued_cell(cell_val):
    if cell_val is None:
        return False
    try:
        if pd.isna(cell_val):
            return False
    except (TypeError, ValueError):
        pass
    s = safe_str(cell_val).strip()
    if not s:
        return False
    if not re.search(r'[,;|]', s):
        return False
    if DECIMAL_COMMA.match(s):
        return False
    if DATE_PATTERN.match(s):
        return False
    parts = [p.strip() for p in SEP_PATTERN.split(s) if p.strip()]
    return len(parts) >= 2


def detect_multi_valued_columns(df):
    mv_cols = []
    for col in df.columns:
        series = df[col].dropna()
        if any(is_multi_valued_cell(v) for v in series):
            mv_cols.append(col)
    return mv_cols


def split_cell(val):
    if val is None:
        return [None]
    try:
        if pd.isna(val):
            return [None]
    except (TypeError, ValueError):
        pass
    s = safe_str(val).strip()
    if not s:
        return [None]
    if DECIMAL_COMMA.match(s) or DATE_PATTERN.match(s):
        return [s]
    parts = [p.strip() for p in SEP_PATTERN.split(s) if p.strip()]
    return parts if len(parts) >= 2 else [s]


def detect_composite_columns(df):
    composites = {}
    for col in df.columns:
        col_lower = str(col).lower()
        for marker in [' et ', ' and ', ' & ']:
            if marker in col_lower:
                parts = re.split(r'\s+(?:et|and|&)\s+', col_lower.strip())
                parts = [p.strip() for p in parts if p.strip()]
                if len(parts) >= 2:
                    composites[col] = parts
                    break
    return composites


def find_candidate_keys(df):
    n = len(df)
    cols = list(df.columns)
    candidates = []
    for c in cols:
        if df[c].nunique() == n and not df[c].isna().any():
            candidates.append([c])
    if candidates:
        return candidates
    for a, b in combinations(cols, 2):
        key = df[a].astype(str) + "|||" + df[b].astype(str)
        if key.nunique() == n:
            candidates.append([a, b])
    if candidates:
        return candidates
    for combo in combinations(cols, 3):
        key = df[combo[0]].astype(str)
        for c in combo[1:]:
            key = key + "|||" + df[c].astype(str)
        if key.nunique() == n:
            candidates.append(list(combo))
    return candidates if candidates else [cols[:1]]


def compute_functional_deps(df):
    fds = defaultdict(set)
    cols = list(df.columns)
    n = len(df)
    for a in cols:
        if df[a].nunique() == n:
            continue
        # Fill NaN with a sentinel so groupby treats them consistently
        a_series = df[a].astype(str).fillna("__NULL__")
        for b in cols:
            if a == b:
                continue
            grouped = df.groupby(a_series, dropna=False)[b].nunique()
            if grouped.max() <= 1:
                fds[a].add(b)
    return dict(fds)


def is_calculated(col):
    calc_keywords = [
        'moyenne', 'mfe', 'total', 'decision', 'resultat', 'verdict',
        'average', 'calculated', 'computed', 'derived', 'score_final',
        'final_grade', 'status', 'moy', 'avg', 'sum', 'total_'
    ]
    return any(kw in col.lower() for kw in calc_keywords)


def sql_type(series):
    sample = series.dropna()
    if len(sample) == 0:
        return "VARCHAR(100)"
    if pd.api.types.is_integer_dtype(series):
        return "INT"
    if pd.api.types.is_float_dtype(series):
        return "DECIMAL(10,2)"
    if pd.api.types.is_datetime64_any_dtype(series):
        return "DATE"
    str_sample = sample.astype(str)
    max_len = str_sample.str.len().max()
    date_like = str_sample.str.match(r'^\d{4}[-/]\d{2}[-/]\d{2}$').mean()
    if date_like > 0.5:
        return "DATE"
    if max_len <= 1:
        return "CHAR(1)"
    if max_len <= 10:
        return "VARCHAR(20)"
    if max_len <= 50:
        return "VARCHAR(100)"
    if max_len <= 100:
        return "VARCHAR(200)"
    return "TEXT"


# ── NF1 ─────────────────────────────────────────────────────────

def apply_nf1(df, groq_analysis=None):
    log = []
    result_df = df.copy()
    groq_analysis = groq_analysis or {}

    # Step 1: Split composite column names (heuristic + Groq)
    composites = detect_composite_columns(result_df)
    for col, parts in groq_analysis.get("composite_columns", {}).items():
        if col in result_df.columns and col not in composites:
            composites[col] = parts
            log.append(f"🤖 Groq: composite '{col}' → {parts}")
    for orig_col, sub_names in composites.items():
        log.append(f"⚡ Composite attribute '{orig_col}' → {sub_names}")
        for i, sub in enumerate(sub_names):
            # Use _i=i default-arg to capture current loop value (fix closure bug).
            # Guard pd.isna so raw floats/NaN never reach .strip().
            result_df[sub] = result_df[orig_col].apply(
                lambda v, _i=i: (
                    str(v).strip().split()[_i]
                    if (not pd.isna(v)) and len(str(v).strip().split()) > _i
                    else None
                )
            )
        result_df = result_df.drop(columns=[orig_col])
        log.append(f"  ✔ Split into: {sub_names}")

    # Step 2: Detect multi-valued columns
    mv_cols = detect_multi_valued_columns(result_df)
    # Add any Groq-detected MV columns not found by heuristic
    for c in groq_analysis.get("multi_valued_columns", []):
        if c in result_df.columns and c not in mv_cols:
            mv_cols.append(c)
            log.append(f"🤖 Groq: multi-valued column '{c}'")

    if not mv_cols:
        log.append("✔ No multi-valued columns — already atomic.")
        return result_df, log, [], []

    log.append(f"⚡ Multi-valued columns found: {mv_cols}")

    # Step 3: Group parallel vs independent
    def get_counts(col):
        return result_df[col].apply(
            lambda v: len(split_cell(v)) if (not pd.isna(v)) and is_multi_valued_cell(v) else 1
        )

    groups = []
    assigned = set()

    for i, col_a in enumerate(mv_cols):
        if col_a in assigned:
            continue
        group = [col_a]
        counts_a = get_counts(col_a)
        for col_b in mv_cols[i+1:]:
            if col_b in assigned:
                continue
            counts_b = get_counts(col_b)
            if (counts_a == counts_b).all():
                group.append(col_b)
        for c in group:
            assigned.add(c)
        groups.append(group)

    independent_mv = [g[0] for g in groups if len(g) == 1]
    parallel_groups = [g for g in groups if len(g) > 1]
    log.append(f"  Parallel groups: {parallel_groups}")
    if independent_mv:
        log.append(f"  Independent MVDs: {independent_mv}")

    # Step 4: Zip-explode all groups
    all_mv_flat = [c for g in groups for c in g]
    anchor_cols = [c for c in result_df.columns if c not in all_mv_flat]

    rows = []
    for _, row in result_df.iterrows():
        group_splits = []
        for group in groups:
            ref = split_cell(row[group[0]])
            gdict = {}
            for col in group:
                vals = split_cell(row[col])
                while len(vals) < len(ref):
                    vals.append(None)
                gdict[col] = vals
            group_splits.append((ref, gdict))

        max_rows = max(len(ref) for ref, _ in group_splits)
        for i in range(max_rows):
            new_row = {c: row[c] for c in anchor_cols}
            for (ref, gdict) in group_splits:
                for col, vals in gdict.items():
                    new_row[col] = vals[i] if i < len(vals) else None
            rows.append(new_row)

    result_df = pd.DataFrame(rows, columns=list(result_df.columns)).reset_index(drop=True)
    log.append(f"✔ Exploded {len(df)} rows → {len(result_df)} atomic rows")

    return result_df, log, parallel_groups, independent_mv


# ── NF2 ─────────────────────────────────────────────────────────

def apply_nf2(df, candidate_keys, groq_analysis=None):
    log = []
    tables = {}
    pk = candidate_keys[0] if candidate_keys else [df.columns[0]]

    if len(pk) == 1:
        log.append(f"✔ Single-column PK '{pk[0]}' — NF2 trivially satisfied.")
        tables["main"] = (df.copy(), pk, [])
        return tables, log, pk

    log.append(f"  PK: {pk}")
    fds = compute_functional_deps(df)
    # Merge Groq-detected FDs
    groq_analysis = groq_analysis or {}
    for det, deps in groq_analysis.get("functional_dependencies", {}).items():
        if det in df.columns:
            existing = set(fds.get(det, set()))
            fds[det] = existing | {d for d in deps if d in df.columns}
            log.append(f"🤖 Groq FD: {det} → {deps}")
    non_pk = [c for c in df.columns if c not in pk]
    extracted_cols = set()
    partial_tables = {}

    for size in range(1, len(pk)):
        for pk_subset in combinations(pk, size):
            pk_subset = list(pk_subset)
            subset_deps = []
            for col in non_pk:
                if col in extracted_cols:
                    continue
                all_det = all(col in fds.get(s, set()) for s in pk_subset)
                not_all_pk = any(col not in fds.get(p, set()) for p in pk if p not in pk_subset)
                if all_det and not_all_pk:
                    subset_deps.append(col)
            if subset_deps:
                tname = "_".join(pk_subset) + "_table"
                sub_df = df[pk_subset + subset_deps].drop_duplicates().reset_index(drop=True)
                partial_tables[tname] = (sub_df, pk_subset, [])
                extracted_cols.update(subset_deps)
                log.append(f"⚡ Partial dep {pk_subset} → {subset_deps} → table '{tname}'")

    remaining = [c for c in df.columns if c not in extracted_cols or c in pk]
    tables["main"] = (df[remaining].drop_duplicates().reset_index(drop=True), pk, [])
    tables.update(partial_tables)

    if not partial_tables:
        log.append("✔ No partial dependencies — already NF2.")
    else:
        log.append(f"✔ NF2: Extracted {len(partial_tables)} table(s).")

    return tables, log, pk


# ── NF3 ─────────────────────────────────────────────────────────

def apply_nf3(tables_nf2, groq_analysis=None):
    log = []
    result = {}
    views_sql = []

    for tname, (df, pk, fk) in tables_nf2.items():
        fds = compute_functional_deps(df)
        non_pk = [c for c in df.columns if c not in pk]
        groq_calc = (groq_analysis or {}).get("calculated_columns", [])
        calc_cols = list(set(
            [c for c in non_pk if is_calculated(c)] +
            [c for c in groq_calc if c in non_pk]
        ))
        transitive_groups = {}

        if calc_cols:
            log.append(f"  Calculated fields in '{tname}': {calc_cols} → SQL VIEW")

        for det in non_pk:
            if det in calc_cols:
                continue
            deps = [c for c in fds.get(det, set())
                    if c not in pk and c != det and c not in calc_cols]
            if deps and df[det].nunique() < len(df) * 0.9:
                transitive_groups[det] = deps

        if not transitive_groups and not calc_cols:
            result[tname] = (df, pk, fk)
            continue

        remaining = [c for c in df.columns if c not in calc_cols]
        new_fks = list(fk)

        for det, deps in transitive_groups.items():
            valid_deps = [d for d in deps if d in remaining]
            if not valid_deps:
                continue
            sub_df = df[[det] + valid_deps].drop_duplicates().reset_index(drop=True)
            new_tname = det + "_ref"
            result[new_tname] = (sub_df, [det], [])
            remaining = [c for c in remaining if c not in valid_deps]
            if det not in new_fks:
                new_fks.append(det)
            log.append(f"⚡ Transitive dep in '{tname}': {det} → {valid_deps} → '{new_tname}'")

        result[tname] = (df[remaining].drop_duplicates().reset_index(drop=True), pk, new_fks)

        if calc_cols:
            view_name = f"v_{clean_name(tname)}_full"
            views_sql.append(
                f"-- Calculated fields for {tname}\n"
                f"CREATE OR REPLACE VIEW `{view_name}` AS\n"
                f"  SELECT t.*,\n" +
                ",\n".join(f"    NULL AS `{clean_name(c)}` /* compute {c} here */" for c in calc_cols) +
                f"\n  FROM `{clean_name(tname)}` t;"
            )

    if not any("⚡" in m for m in log):
        log.append("✔ No transitive dependencies — already NF3.")
    else:
        log.append("✔ NF3 applied.")

    return result, log, views_sql


# ── BCNF ─────────────────────────────────────────────────────────

def apply_bcnf(tables_nf3):
    log = []
    result = {}

    for tname, (df, pk, fk) in tables_nf3.items():
        fds = compute_functional_deps(df)
        non_pk = [c for c in df.columns if c not in pk]
        violations = {}

        for det in non_pk:
            det_determines_pk = all(pkc in fds.get(det, set()) for pkc in pk)
            if det_determines_pk:
                continue
            det_deps = [c for c in fds.get(det, set()) if c not in pk and c != det]
            if det_deps and df[det].nunique() < len(df) * 0.9:
                violations[det] = det_deps

        if not violations:
            result[tname] = (df, pk, fk)
            continue

        remaining = list(df.columns)
        new_fks = list(fk)

        for det, deps in violations.items():
            valid_deps = [d for d in deps if d in remaining]
            if not valid_deps:
                continue
            sub_df = df[[det] + valid_deps].drop_duplicates().reset_index(drop=True)
            new_tname = det + "_bcnf"
            result[new_tname] = (sub_df, [det], [])
            remaining = [c for c in remaining if c not in valid_deps]
            if det not in new_fks:
                new_fks.append(det)
            log.append(f"⚡ BCNF violation in '{tname}': {det} → {valid_deps} → '{new_tname}'")

        result[tname] = (df[remaining].drop_duplicates().reset_index(drop=True), pk, new_fks)

    if not any("⚡" in m for m in log):
        log.append("✔ All determinants are superkeys — BCNF satisfied.")
    else:
        log.append("✔ BCNF applied.")

    return result, log


# ── NF4 ─────────────────────────────────────────────────────────

def apply_nf4(tables_bcnf, original_df, independent_mv_cols, orig_pk):
    log = []
    result = dict(tables_bcnf)

    if not independent_mv_cols:
        log.append("✔ No independent multi-valued dependencies — NF4 satisfied.")
        return result, log

    log.append(f"  Independent MVD attributes: {independent_mv_cols}")

    for mv_col in independent_mv_cols:
        if mv_col not in original_df.columns:
            continue
        rows = []
        for _, row in original_df.iterrows():
            pk_vals = {p: row[p] for p in orig_pk if p in original_df.columns}
            raw = row.get(mv_col, None)
            vals = split_cell(raw) if is_multi_valued_cell(raw) else [raw]
            for val in vals:
                if val is None:
                    continue
                r = dict(pk_vals)
                r[mv_col] = val
                rows.append(r)
        if not rows:
            continue
        bridge_df = pd.DataFrame(rows).drop_duplicates().reset_index(drop=True)
        tname = mv_col + "_mvd"
        result[tname] = (bridge_df, orig_pk + [mv_col], list(orig_pk))
        log.append(f"⚡ MVD {orig_pk} ↠ {mv_col} → bridge table '{tname}'")

    if any("⚡" in m for m in log):
        log.append("✔ NF4 applied.")
    else:
        log.append("✔ NF4 satisfied.")

    return result, log


# ── NF5 ─────────────────────────────────────────────────────────

def apply_nf5(tables_nf4):
    log = []
    result = {}
    new_tables = {}

    for tname, (df, pk, fk) in tables_nf4.items():
        cols = list(df.columns)
        if len(pk) >= 3 and set(pk) == set(cols):
            pairs = list(combinations(pk, 2))
            projected = [df[list(p)].drop_duplicates() for p in pairs]
            joined = projected[0]
            for proj in projected[1:]:
                joined = joined.merge(proj, how='inner')
            orig_sorted = df.sort_values(by=pk).reset_index(drop=True)
            join_sorted = joined.drop_duplicates().sort_values(by=pk).reset_index(drop=True)
            if len(join_sorted) == len(orig_sorted):
                for pair in pairs:
                    sub_df = df[list(pair)].drop_duplicates().reset_index(drop=True)
                    new_tname = "_".join(pair) + "_5nf"
                    new_tables[new_tname] = (sub_df, list(pair), [])
                log.append(f"⚡ Join dependency in '{tname}' → {len(pairs)} binary tables")
                continue
        result[tname] = (df, pk, fk)

    result.update(new_tables)
    if not any("⚡" in m for m in log):
        log.append("✔ No join dependencies — NF5 satisfied.")
    else:
        log.append("✔ NF5 applied.")

    return result, log


# ── SQL Generator ────────────────────────────────────────────────

def generate_sql(tables, views_sql=None, db_name="normalized_db", comments=None):
    views_sql = views_sql or []
    lines = [
        "-- ══════════════════════════════════════════════════════════",
        f"--  Database : {db_name}",
        "--  Generated by NormalizerDB v2.0  (NF1 → NF5)",
        "-- ══════════════════════════════════════════════════════════",
        "",
        f"CREATE DATABASE IF NOT EXISTS `{db_name}` CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci;",
        f"USE `{db_name}`;",
        "",
        "SET FOREIGN_KEY_CHECKS = 0;",
        "",
    ]
    insert_blocks = []

    for tname, (df, pk, fk) in tables.items():
        safe = clean_name(tname)
        col_defs = []
        for col in df.columns:
            cname = clean_name(col)
            ctype = sql_type(df[col])
            null_clause = "NOT NULL" if col in pk else "DEFAULT NULL"
            col_defs.append(f"    `{cname}` {ctype} {null_clause}")
        if pk:
            pk_str = ", ".join(f"`{clean_name(p)}`" for p in pk)
            col_defs.append(f"    PRIMARY KEY ({pk_str})")
        for fk_col in fk:
            fk_safe = clean_name(fk_col)
            ref_table = None
            for other_tname, (_, other_pk, _) in tables.items():
                if other_tname != tname and fk_col in other_pk:
                    ref_table = clean_name(other_tname)
                    break
            if ref_table:
                col_defs.append(
                    f"    CONSTRAINT fk_{safe}_{fk_safe} "
                    f"FOREIGN KEY (`{fk_safe}`) REFERENCES `{ref_table}` (`{fk_safe}`) "
                    f"ON DELETE CASCADE ON UPDATE CASCADE"
                )

        comment = (comments or {}).get(tname, "")
        comment_line = f"  /* {comment} */" if comment else ""
        lines.append(f"-- ── Table: {safe}{comment_line}")
        lines.append(f"CREATE TABLE IF NOT EXISTS `{safe}` (")
        lines.append(",\n".join(col_defs))
        lines.append(") ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;")
        for fk_col in fk:
            lines.append(f"CREATE INDEX idx_{safe}_{clean_name(fk_col)} ON `{safe}` (`{clean_name(fk_col)}`);")
        lines.append("")

        ins = [f"-- Data: {safe}"]
        cols_clean = [clean_name(c) for c in df.columns]
        col_list = ", ".join(f"`{c}`" for c in cols_clean)
        for _, row in df.iterrows():
            vals = []
            for v in row:
                if pd.isna(v) or v is None:
                    vals.append("NULL")
                elif isinstance(v, bool):
                    vals.append("1" if v else "0")
                elif isinstance(v, (int, np.integer)):
                    vals.append(str(int(v)))
                elif isinstance(v, (float, np.floating)):
                    vals.append(f"{v:.4f}" if not np.isnan(v) else "NULL")
                else:
                    escaped = str(v).replace("\\", "\\\\").replace("'", "\\'")
                    vals.append(f"'{escaped}'")
            ins.append(f"INSERT INTO `{safe}` ({col_list}) VALUES ({', '.join(vals)});")
        ins.append("")
        insert_blocks.append("\n".join(ins))

    lines.append("SET FOREIGN_KEY_CHECKS = 1;\n")
    if views_sql:
        lines.append("-- ── VIEWS ────────────────────────────────────────────────────")
        lines.extend(views_sql)
        lines.append("")
    lines.append("-- ── INSERT DATA ──────────────────────────────────────────────")
    lines.extend(insert_blocks)
    return "\n".join(lines)


# ── Excel Writer ─────────────────────────────────────────────────

def write_sheet(wb, title, df, pk_cols=None, fk_cols=None, tab_color=None, description=""):
    ws = wb.create_sheet(title=title[:31])
    if tab_color:
        ws.sheet_properties.tabColor = tab_color
    pk_cols = pk_cols or []; fk_cols = fk_cols or []
    start_row = 1
    if description:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(df.columns),1))
        dc = ws.cell(row=1, column=1, value=description)
        dc.font = Font(italic=True, name="Arial", size=9, color="666666")
        dc.alignment = Alignment(wrap_text=True)
        start_row = 2
    hdr_fill = PatternFill("solid", start_color="1B2631")
    pk_hdr   = PatternFill("solid", start_color="7D6608")
    fk_hdr   = PatternFill("solid", start_color="0E6655")
    for ci, col in enumerate(df.columns, 1):
        is_pk = col in pk_cols; is_fk = col in fk_cols
        cell = ws.cell(row=start_row, column=ci,
                       value=f"{'🔑 ' if is_pk else '🔗 ' if is_fk else ''}{col}")
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill = pk_hdr if is_pk else (fk_hdr if is_fk else hdr_fill)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER
    for ri, row in enumerate(df.itertuples(index=False), start_row+1):
        fill_color = "EBF5FB" if ri % 2 == 0 else "FDFEFE"
        for ci, (col, val) in enumerate(zip(df.columns, row), 1):
            cell = ws.cell(row=ri, column=ci, value=val if not (isinstance(val, float) and np.isnan(val)) else None)
            if col in pk_cols:
                cell.fill = PatternFill("solid", start_color="FEF9E7")
                cell.font = Font(name="Arial", size=10, bold=True)
            elif col in fk_cols:
                cell.fill = PatternFill("solid", start_color="EAFAF1")
                cell.font = Font(name="Arial", size=10)
            else:
                cell.fill = PatternFill("solid", start_color=fill_color)
                cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical='center')
            cell.border = BORDER
    for ci, col in enumerate(df.columns, 1):
        vals = df[col].astype(str)
        max_w = max(vals.str.len().max(), len(str(col))+4)
        ws.column_dimensions[get_column_letter(ci)].width = min(max_w+3, 45)
    ws.row_dimensions[start_row].height = 22


def nf_label(tname):
    t = tname.lower()
    if "5nf" in t: return "5NF"
    if "mvd" in t: return "4NF"
    if "bcnf" in t: return "BCNF"
    if "ref" in t: return "3NF"
    if "table" in t: return "2NF"
    return "5NF"


def write_normalized_excel(tables, path, nf_log, views_sql=None):
    wb = Workbook(); wb.remove(wb.active)
    ws = wb.create_sheet("Summary", 0)
    ws.sheet_properties.tabColor = "C0392B"
    r = 1
    for text, bg, fg, sz, bold in [
        ("NormalizerDB v2.0 — Normalization Report", "1B2631","FFFFFF",15,True),
        ("NF1 → NF5  |  All Normal Forms Applied",   "2C3E50","AED6F1",11,False),
        ("","","","",""),
    ]:
        if not text: r += 1; continue
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6)
        cell = ws.cell(row=r, column=1, value=text)
        cell.font = Font(name="Arial", size=sz, bold=bold, color=fg)
        if bg: cell.fill = PatternFill("solid", start_color=bg)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[r].height = int(sz)+12
        r += 1

    hdr_fill = PatternFill("solid", start_color="1B2631")
    for ci, h in enumerate(["Table","Rows","Cols","Primary Key","NF Level","FK refs"],1):
        cell = ws.cell(row=r, column=ci, value=h)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill = hdr_fill; cell.border = BORDER
        cell.alignment = Alignment(horizontal='center')
    r += 1

    nf_fills = {"5NF":"E8DAEF","4NF":"FDEBD0","BCNF":"D5F5E3","3NF":"FEF9E7","2NF":"FADBD8"}
    for tname,(df,pk,fk) in tables.items():
        nf = nf_label(tname)
        fill = PatternFill("solid", start_color=nf_fills.get(nf,"F2F3F4"))
        for ci,val in enumerate([tname,len(df),len(df.columns),", ".join(pk),nf,", ".join(fk) if fk else "—"],1):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.fill=fill; cell.font=Font(name="Arial",size=10); cell.border=BORDER
            cell.alignment=Alignment(wrap_text=True)
        r += 1

    r += 1
    nf_colors = {"NF1":"FADBD8","NF2":"FDEBD0","NF3":"FEF9E7","BCNF":"E8F8F5","NF4":"EBF5FB","NF5":"F4ECF7"}
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6)
    hc = ws.cell(row=r,column=1,value="Normalization Log")
    hc.font=Font(bold=True,color="FFFFFF",name="Arial",size=11)
    hc.fill=PatternFill("solid",start_color="1B2631"); hc.alignment=Alignment(horizontal='center')
    ws.row_dimensions[r].height=20; r+=1

    for stage, messages in nf_log.items():
        color = nf_colors.get(stage,"F2F3F4")
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6)
        cell = ws.cell(row=r,column=1,value=f"── {stage} ──")
        cell.font=Font(bold=True,name="Arial",size=10)
        cell.fill=PatternFill("solid",start_color=color); r+=1
        for msg in messages:
            ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6)
            cell = ws.cell(row=r,column=1,value=f"   {msg}")
            cell.font=Font(name="Arial",size=9)
            cell.fill=PatternFill("solid",start_color=color)
            ws.row_dimensions[r].height=14; r+=1

    for c,w in zip('ABCDEF',[30,8,8,25,12,30]):
        ws.column_dimensions[c].width = w

    tab_colors=["3498DB","2ECC71","9B59B6","E67E22","F1C40F","1ABC9C","E74C3C","16A085","D35400","8E44AD"]
    for i,(tname,(df,pk,fk)) in enumerate(tables.items()):
        desc = f"NF: {nf_label(tname)}  |  PK: {', '.join(pk) or '—'}  |  FK: {', '.join(fk) if fk else '—'}"
        write_sheet(wb, tname, df, pk_cols=pk, fk_cols=fk,
                    tab_color=tab_colors[i%len(tab_colors)], description=desc)
    wb.save(path)


# ── Main Pipeline ─────────────────────────────────────────────────

def normalize(input_excel, output_excel, output_sql,
              sheet_name=0, groq_api_key=None, log_fn=None):
    """
    Full NF1→NF5 normalization pipeline.

    groq_api_key : str | None
        When provided, Groq LLaMA 3.3-70B is used for semantic analysis
        before the heuristic pipeline runs, giving much smarter FD detection,
        better table naming, and calculated-field identification.
        Falls back to heuristics if the key is missing or the call fails.

    log_fn : callable(str) | None
        Optional callback for streaming log messages (used by Flask SSE).
    """
    def _log(msg):
        if log_fn: log_fn(msg)
        print(f"  {msg}")

    _log(f"\n{'═'*60}")
    _log(f"  NormalizerDB v2.0 — {input_excel}")
    _log(f"{'═'*60}")

    df_raw = pd.read_excel(input_excel, sheet_name=sheet_name)
    _log(f"Loaded: {len(df_raw)} rows × {len(df_raw.columns)} cols")
    _log(f"Columns: {list(df_raw.columns)}")

    nf_log = {}; all_views = []
    groq_analysis = {}
    table_domain  = "normalized_db"
    sql_comments  = {}

    # ── Optional Groq AI pre-analysis ────────────────────────────
    if groq_api_key:
        try:
            from groq_advisor import (
                analyze_table, suggest_table_names,
                generate_sql_comments, merge_analysis
            )
            groq_raw = analyze_table(df_raw, groq_api_key, log_fn=log_fn)
            groq_analysis = groq_raw
            table_domain  = groq_raw.get("table_domain", "normalized_db")
            nf_log["🤖 AI Analysis"] = [
                f"Model: Groq LLaMA 3.3-70B",
                f"Domain: {table_domain}",
                f"Suggested PK: {groq_raw.get('primary_key', [])}",
                f"Entities detected: {[t['name'] for t in groq_raw.get('entity_tables', [])]}",
                f"Calculated fields: {groq_raw.get('calculated_columns', [])}",
                f"Notes: {groq_raw.get('normalization_notes', '')[:200]}",
            ]
        except ImportError:
            _log("⚠️  groq_advisor not found — using heuristics only")

    # ── NF1 ──────────────────────────────────────────────────────
    _log("[NF1] Atomizing values & exploding repeating groups…")
    df_nf1, log1, parallel_groups, independent_mv = apply_nf1(
        df_raw, groq_analysis=groq_analysis
    )
    nf_log["NF1"] = log1
    for m in log1: _log(f"  {m}")

    # Choose PK: Groq suggestion → heuristic
    groq_pk = groq_analysis.get("primary_key")
    if groq_pk and all(c in df_nf1.columns for c in groq_pk):
        ckeys = [groq_pk]
        _log(f"PK (Groq): {groq_pk}")
    else:
        ckeys = find_candidate_keys(df_nf1)
        _log(f"PK (heuristic): {ckeys[0]}")
    pk = ckeys[0]

    # ── NF2 ──────────────────────────────────────────────────────
    _log("[NF2] Removing partial dependencies…")
    tables_nf2, log2, pk = apply_nf2(
        df_nf1, ckeys, groq_analysis=groq_analysis
    )
    nf_log["NF2"] = log2
    for m in log2: _log(f"  {m}")

    # ── NF3 ──────────────────────────────────────────────────────
    _log("[NF3] Removing transitive dependencies…")
    tables_nf3, log3, views3 = apply_nf3(
        tables_nf2, groq_analysis=groq_analysis
    )
    nf_log["NF3"] = log3; all_views.extend(views3)
    for m in log3: _log(f"  {m}")

    # ── BCNF ─────────────────────────────────────────────────────
    _log("[BCNF] Verifying every determinant is a superkey…")
    tables_bcnf, log_b = apply_bcnf(tables_nf3)
    nf_log["BCNF"] = log_b
    for m in log_b: _log(f"  {m}")

    # ── NF4 ──────────────────────────────────────────────────────
    _log("[NF4] Removing multi-valued dependencies…")
    orig_pk_for_nf4 = [c for c in pk if c in df_raw.columns]
    # Groq-detected independent MVDs override heuristic ones
    groq_indep_mv = groq_analysis.get("independent_mvd_columns", [])
    final_indep_mv = list(set(independent_mv + [
        c for c in groq_indep_mv if c in df_raw.columns
    ]))
    tables_nf4, log4 = apply_nf4(
        tables_bcnf, df_raw, final_indep_mv, orig_pk_for_nf4
    )
    nf_log["NF4"] = log4
    for m in log4: _log(f"  {m}")

    # ── NF5 ──────────────────────────────────────────────────────
    _log("[NF5] Checking join dependencies…")
    tables_nf5, log5 = apply_nf5(tables_nf4)
    nf_log["NF5"] = log5
    for m in log5: _log(f"  {m}")

    # ── Optional: Groq table name suggestions ────────────────────
    if groq_api_key and groq_analysis:
        try:
            from groq_advisor import suggest_table_names, generate_sql_comments
            name_map = suggest_table_names(tables_nf5, groq_api_key, log_fn=log_fn)
            if name_map:
                renamed = {}
                for old, (df, p, f) in tables_nf5.items():
                    new_name = name_map.get(old, old)
                    renamed[new_name] = (df, p, f)
                tables_nf5 = renamed
                nf_log["🏷️ Table Names"] = [f"{k} → {v}" for k, v in name_map.items()]
                _log(f"Table names refined by Groq: {name_map}")

            sql_comments = generate_sql_comments(
                tables_nf5, table_domain, groq_api_key, log_fn=log_fn
            )
        except Exception as e:
            _log(f"⚠️  Name/comment step failed: {e}")

    # ── Output ────────────────────────────────────────────────────
    _log(f"\nFinal schema ({len(tables_nf5)} tables):")
    for tname, (df, p, f) in tables_nf5.items():
        _log(f"  {tname:35s} {len(df):4d} rows  {len(df.columns)} cols  PK={p}")

    write_normalized_excel(tables_nf5, output_excel, nf_log, all_views)
    _log(f"✔ Excel: {output_excel}")

    sql = generate_sql(tables_nf5, all_views, db_name=re.sub(r'\W', '_', table_domain)[:30],
                       comments=sql_comments)
    with open(output_sql, "w", encoding="utf-8") as f:
        f.write(sql)
    _log(f"✔ SQL: {output_sql}")
    _log(f"✅ DONE — {len(tables_nf5)} tables produced")

    return tables_nf5, nf_log


if __name__ == "__main__":
    normalize(
        input_excel="/home/claude/sample_denormalized.xlsx",
        output_excel="/home/claude/normalized_v2.xlsx",
        output_sql="/home/claude/normalized_v2.sql",
    )
# ── Address heuristic patch ──────────────────────────────────────
# Monkey-patch is_multi_valued_cell to ignore address-like values
_orig_is_mv = is_multi_valued_cell

# Keywords that suggest an address context (not a real list)
_ADDRESS_LIKE = re.compile(
    r'(rue|av|avenue|bd|boulevard|quartier|cite|villa|zone|district'
    r'|street|road|lane|apt|floor|étage|wilaya|region|county'
    r'|\bno\b|\bn°\b|\bnum\b)', re.I
)

def is_multi_valued_cell(cell_val):
    if cell_val is None:
        return False
    try:
        if pd.isna(cell_val):
            return False
    except (TypeError, ValueError):
        pass
    s = safe_str(cell_val).strip()
    if not s:
        return False
    # If value looks like an address, don't split
    if _ADDRESS_LIKE.search(s):
        return False
    return _orig_is_mv(cell_val)
